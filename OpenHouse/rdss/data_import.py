from openpyxl import load_workbook
from rdss.models import Student
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def ImportStudentCardID(file):
    wb = load_workbook(file)
    sheet = wb.worksheets[0]

    for i in range(2, sheet.max_row+1):
        Student.objects.update_or_create(
            idcard_no= sheet.cell(row=i, column=2).value,
            defaults={
                'student_id': sheet.cell(row=i, column=1).value or '',
                'name': sheet.cell(row=i, column=3).value or '',
                'email': sheet.cell(row=i, column=4).value or '',
                'phone': sheet.cell(row=i, column=5).value or '',
                'dep': sheet.cell(row=i, column=6).value or '',
            }
        )
