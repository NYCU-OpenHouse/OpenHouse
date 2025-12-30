from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, JsonResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from company.forms import CompanyCreationForm, CompanyEditForm, CompanyPasswordResetForm
from django.utils.encoding import force_text
from django.utils.http import urlsafe_base64_decode
from .models import Company, ChineseFundedCompany, Job
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.forms import SetPasswordForm
from company.forms import CompanyCreationForm, CompanyEditForm, ItemModelFormSet
from oauth2_provider.decorators import protected_resource
from oauth.models import CustomAccessToken
from django.http import Http404
from openpyxl import load_workbook

import rdss.models
import recruit.models
import company.models
import general.models
import recruitment_common.views as views_helper


@login_required(login_url='/company/login/')
def CompanyIndex(request):
    # semantic ui control
    menu_ui = {'index': "active"}

    # rdss files
    rdss_file_list = rdss.models.Files.objects.all().order_by('-updated')
    recruit_file_list = recruit.models.Files.objects.all().order_by('-updated')
    show_recruit = general.models.NavbarConfigs.objects.first().show_recruit
    show_rdss = general.models.NavbarConfigs.objects.first().show_rdss
    return render(request, 'company_index.html', locals())


@login_required(login_url='/company/login/')
def CompanyInfo(request):
    # semantic ui control
    menu_ui = {'info': "active"}

    company_info = company.models.Company.objects.get(cid=request.user.cid)
    jobs = company.models.Job.objects.filter(cid=request.user.id)
    return render(request, 'company_info.html', locals())


def CompanyCreation(request):
    submit_btn_name = "創建帳號"
    if request.POST:
        form = CompanyCreationForm(request.POST, request.FILES)
        try:
            ChineseFundedCompany.objects.get(cid=request.POST['cid'])
            error_display = True
            error_msg = "此統編已被政府列為中資企業，若有註冊需求，歡迎來信詢求協助!"
            return render(request, 'company_create_form.html', locals())
        except ChineseFundedCompany.DoesNotExist:
            pass
        if form.is_valid():
            form.save()
            user = authenticate(username=form.clean_cid(), password=form.clean_password2())
            login(request, user)
            messages.success(request, '公司帳號創建成功！請至編輯頁最下方新增公司招募職缺')
            return redirect(CompanyEdit)
        else:
            error_display = True
            error_msg = form.errors
            return render(request, 'company_create_form.html', locals())
    form = CompanyCreationForm()
    return render(request, 'company_create_form.html', locals())

def _truncate_unicode_by_bytes(input_string, max_bytes):
    encoded_string = input_string.encode('utf-8')
    if len(encoded_string) <= max_bytes:
        return input_string

    truncated_bytes = encoded_string[:max_bytes]
    while True:
        try:
            truncated_string = truncated_bytes.decode('utf-8')
            return truncated_string
        except UnicodeDecodeError:
            truncated_bytes = truncated_bytes[:-1]


def _update_job_position_with_excel_file(excel_file, company: Company):
    MAX_DESCRIPTION_BYTES = 780
    wb = load_workbook(excel_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):

        if not any(row):
            continue
        (title, quantity, is_liberal, is_foreign, description, 
        education, salary, welfare, vacation, note, 
        english_title, english_description,
        english_education, english_salary, english_welfare, english_vacation, english_note) = row
        
        description = _truncate_unicode_by_bytes(description, MAX_DESCRIPTION_BYTES) if description is not None else ""
        education = _truncate_unicode_by_bytes(education, MAX_DESCRIPTION_BYTES) if education is not None else ""
        salary = _truncate_unicode_by_bytes(salary, MAX_DESCRIPTION_BYTES) if salary is not None else ""
        welfare = _truncate_unicode_by_bytes(welfare, MAX_DESCRIPTION_BYTES) if welfare is not None else ""
        vacation = _truncate_unicode_by_bytes(vacation, MAX_DESCRIPTION_BYTES) if vacation is not None else ""
        note = _truncate_unicode_by_bytes(note, MAX_DESCRIPTION_BYTES) if note is not None else ""
        english_title = english_title if english_title is not None else ""
        english_description = _truncate_unicode_by_bytes(english_description, MAX_DESCRIPTION_BYTES) if english_description is not None else ""
        english_education = _truncate_unicode_by_bytes(english_education, MAX_DESCRIPTION_BYTES) if english_education is not None else ""
        english_salary = _truncate_unicode_by_bytes(english_salary, MAX_DESCRIPTION_BYTES) if english_salary is not None else ""
        english_welfare = _truncate_unicode_by_bytes(english_welfare, MAX_DESCRIPTION_BYTES) if english_welfare is not None else ""
        english_vacation = _truncate_unicode_by_bytes(english_vacation, MAX_DESCRIPTION_BYTES) if english_vacation is not None else ""
        english_note = _truncate_unicode_by_bytes(english_note, MAX_DESCRIPTION_BYTES) if english_note is not None else ""

        Job.objects.create(
            cid=company,
            title=title,
            is_liberal=is_liberal,
            is_foreign=is_foreign,
            description=description,
            education=education,
            salary=salary,
            welfare=welfare,
            vacation=vacation,
            quantity=quantity,
            note=note,
            english_title=english_title,
            english_description=english_description,
            english_education=english_education,
            english_salary=english_salary,
            english_welfare=english_welfare,
            english_vacation=english_vacation,
            english_note=english_note
        )


@login_required(login_url='/company/login/')
def CompanyEdit(request):

    submit_btn_name = "確認修改"
    if request.user and request.user.is_authenticated:
        user = request.user
    else:
        user = None
        
    company_info = company.models.Company.objects.get(cid=request.user.cid)
    jobs = Job.objects.filter(cid=request.user.id)

    if request.POST:
        form = CompanyEditForm(request.POST, request.FILES, instance=user)
        
        if form.is_valid():
            user = form.save()
            job_formset = ItemModelFormSet(request.POST, instance=user)
            if job_formset.is_valid():
                job_formset.save()
            else:
                messages.error(request, "「職缺列表」填入資料有誤，請重新修改公司資料（錯誤欄位會有提示字眼，請下滑檢查）。")
                return render(request, 'company_edit_form.html', locals())
            
            excel_file = request.FILES.get('excel_file')
            if excel_file:
                try:
                    _update_job_position_with_excel_file(excel_file, user)
                    messages.success(request, '「職缺上傳檔案」成功，請確認內容並再次送出修改公司資料，注意中英文欄位「職缺內容、備註」只會存取前260個字元')
                    return redirect('company_edit')
                
                except Exception as e:
                    print(f"Error during job creation: {e}")
                    messages.error(request, f"「職缺上傳檔案」填入資料有誤，請注意欄位「職缺名稱、職缺數量、是否為文組職缺、是否開放外籍生投遞、職缺內容為必填欄位」。錯誤內容：{e}")
                    return render(request, 'company_edit_form.html', locals())
                
            return redirect(CompanyInfo)
        else:
            messages.error(request, f"「公司資料」填入資料有誤，請重新修改公司資料。錯誤內容：{form.errors}")
            return render(request, 'company_edit_form.html', locals())
    else:
        form = CompanyEditForm(instance=user)
        job_formset = ItemModelFormSet(instance=user)
        job_formset.extra = 1 if job_formset.queryset.count() < 1 else 0
    
    if (len(jobs) == 0):
        messages.info(request, '目前沒有招募職缺，請新增公司招募職缺') 
    if (company_info.categories.name == '其他'):
        messages.error(request, '目前公司類別為其他，請修改公司類別')

    return render(request, 'company_edit_form.html', locals())


def CompanyLogin(request):
    if request.POST:
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
            next_url = request.GET.get('next')
            if user.is_staff:
                if next_url == None:
                    return redirect('/admin/')
                else:
                    return redirect(next_url)
            else:
                if next_url == None:
                    return redirect('/company/')
                else:
                    return redirect(next_url)
        else:
            error_display = True
            error_msg = "帳號或密碼錯誤"

    return render(request, 'login.html', locals())


def CompanyLogout(request):
    logout(request)
    return redirect('/company/login/')


def forget_password(request):
    send = None
    if request.method == 'POST':
        data = request.POST.copy()
        data['email'] = "dummy@e.mail"
        form = CompanyPasswordResetForm(data)
        if form.is_valid():
            form.save(request=request)
            company_obj = company.models.Company.objects.get(cid=form.cleaned_data.get('user'))
            return render(request, 'notify_password_reset.html', locals())
    else:
        form = CompanyPasswordResetForm()
    return render(request, 'forget_password.html', {'form': form, 'send': send})


def password_reset_confirm(request, uidb64, token):
    try:
        uid = force_text(urlsafe_base64_decode(uidb64))
        user = Company.objects.get(pk=uid)
    except(TypeError):
        user = None

    # the link is valid
    if user is not None and default_token_generator.check_token(user, token):
        validlink = True
        if request.method == 'POST':
            form = SetPasswordForm(user, request.POST)
            if form.is_valid():
                form.save()
                return redirect('/company/login/')
        else:
            form = SetPasswordForm(user)
            return render(request, 'password_reset_confirm.html', {'form': form})

    # the link is not valid
    else:
        validlink = False
    return redirect('/')


def ResetPassword(request):
    user = Company.objects.get(cid=request.user.cid)
    if request.method == 'POST':
        form = SetPasswordForm(user, request.POST)
        if form.is_valid():
            form.save()
            return redirect(CompanyInfo)
    form = SetPasswordForm(user)
    return render(request, 'password_reset_confirm.html', {'form': form})

@protected_resource()
def get_company_id(request):

    access_token = CustomAccessToken.objects.get(token=request.META.get('HTTP_AUTHORIZATION').split(' ')[1])

    data = {
        "company_id": access_token.user.cid,
    }
    return JsonResponse(data)

@staff_member_required
def regitered_chinese_funded_company(request):
    
    registered_chinese_funded_company = []
    
    companies = Company.objects.all()
    
    for company in companies:
        if len(ChineseFundedCompany.objects.filter(cid=company.cid)) == 1:
            registered_chinese_funded_company.append(company)
    
    
    return render(request, 'admin/registered_chinese_funded_company.html', locals())

def CompanyDetail(request, companyId):
    """
    Function displaying specific company info to public
    """
    try:
        company_info = Company.objects.get(cid=companyId)
    except Company.DoesNotExist:
        raise Http404("Company not found")
    company_info.website = views_helper.change_website_start_with_http(company_info.website)
    id = company_info.id
    jobs = company.models.Job.objects.filter(cid=id)

    return render(request, 'company_detail.html', locals())
